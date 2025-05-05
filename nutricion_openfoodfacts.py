import openpyxl
import requests
import logging
import time

# CONFIGURACIÓN DE LOG
logging.basicConfig(
    filename="log_alimentos_openfoodfacts.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# FUNCIÓN PARA CONSULTAR LA API DE OPENFOODFACTS
def buscar_info_nutricional(nombre_alimento):
    url = "https://world.openfoodfacts.org/cgi/search.pl"
    params = {
        "search_terms": nombre_alimento,
        "search_simple": 1,
        "action": "process",
        "json": 1,
        "page_size": 1
    }
    try:
        response = requests.get(url, params=params)
        response.raise_for_status()
        datos = response.json()
        productos = datos.get("products", [])
        if not productos:
            return None

        producto = productos[0]
        nutriments = producto.get("nutriments", {})
        parte_comestible = producto.get("completeness", "")  # puede no estar
        return {
            "calorias": nutriments.get("energy-kcal_100g"),
            "proteinas": nutriments.get("proteins_100g"),
            "grasas": nutriments.get("fat_100g"),
            "hidratos": nutriments.get("carbohydrates_100g"),
            "azucares": nutriments.get("sugars_100g"),
            "fibra": nutriments.get("fiber_100g"),
            "sal": nutriments.get("salt_100g"),
            "parte_comestible": parte_comestible
        }
    except Exception as e:
        logging.error(f"Error consultando '{nombre_alimento}': {e}")
        return None

# CARGAR EXCEL
ruta_excel = r"C:\Users\Jaime\OneDrive - Medios Audiovisuales Masmedia SL\jj\scripts\bedca_scraping\alimentos.xlsx"
wb = openpyxl.load_workbook(ruta_excel)
ws = wb.active

# CABECERAS NUEVAS
cabeceras = [
    "Calorías (kcal/100g)", "Proteínas (g/100g)", "Grasas (g/100g)",
    "Hidratos (g/100g)", "Azúcares (g/100g)", "Fibra (g/100g)",
    "Sal (g/100g)", "Parte comestible (%)"
]

# AÑADIR CABECERAS NUEVAS SI NO ESTÁN
if ws.max_column < 2:
    for idx, titulo in enumerate(cabeceras, start=2):
        ws.cell(row=1, column=idx).value = titulo

# PROCESAR ALIMENTOS
for row in ws.iter_rows(min_row=2, max_col=1):
    celda = row[0]
    nombre = celda.value
    if not nombre:
        continue

    logging.info(f"Buscando: {nombre}")
    datos = buscar_info_nutricional(nombre)
    if datos:
        ws.cell(row=celda.row, column=2).value = datos.get("calorias")
        ws.cell(row=celda.row, column=3).value = datos.get("proteinas")
        ws.cell(row=celda.row, column=4).value = datos.get("grasas")
        ws.cell(row=celda.row, column=5).value = datos.get("hidratos")
        ws.cell(row=celda.row, column=6).value = datos.get("azucares")
        ws.cell(row=celda.row, column=7).value = datos.get("fibra")
        ws.cell(row=celda.row, column=8).value = datos.get("sal")
        ws.cell(row=celda.row, column=9).value = datos.get("parte_comestible")
        logging.info(f"✔ Datos encontrados para '{nombre}'")
    else:
        logging.warning(f"✖ No se encontraron datos para '{nombre}'")

    time.sleep(1.2)  # Evitar sobrecargar la API

# GUARDAR RESULTADOS
ruta_salida = r"C:\Users\Jaime\OneDrive - Medios Audiovisuales Masmedia SL\jj\scripts\bedca_scraping\alimentos_completado.xlsx"
wb.save(ruta_salida)
logging.info("✔ Archivo completado guardado con éxito.")

print(f"✅ Proceso finalizado. Archivo guardado en: {ruta_salida}")
